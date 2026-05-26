"""Transforma o relatório 'Contas a Pagar - Documentos em Aberto' (XLS HTML do GCOM)
no template de importação de Títulos a Pagar (aba 'Preencher').
"""
import io
import os
import re
from typing import Optional

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "template",
    "TEMPLATE_TITULOS_A_PAGAR_(GCOM).xlsx",
)

# Termos que identificam que a tabela é a de Contas a Pagar.
HEADER_SIGNATURES = [
    "VENCTO",
    "DATA DA ENTRADA",
    "N DOCTO",
    "Nº DOCTO",
    "CNPJ/CPF",
    "FORNECEDOR",
]

# Posições 1-based no template (linha 1 = cabeçalho).
COLUNAS_FIXAS = {
    1:  "PP",     # A  - Identificação do tipo de integração de título
    5:  "0001",   # E  - Empresa Emitente
    6:  "0001",   # F  - Código da Filial
    7:  "0001",   # G  - Empresa Pagadora
    8:  55,       # H  - Tipo de Título
    12: "BRL",    # L  - Código da Moeda
    13: "CA",     # M  - Tipo de Cobrança
}

VALOR_GRUPO_PAGAMENTO = 1106010000
VALOR_FLUXO_CAIXA = "01"


def _formatar_data(data_str) -> str:
    """'05/01/2026' -> '05012026'. Retorna '' se vazio/None/NaN."""
    if data_str is None:
        return ""
    s = str(data_str).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    return s.replace("/", "")


def _limpar_documento(doc) -> str:
    """Remove pontuação de CNPJ/CPF, mantém só dígitos.
    CNPJ esperado 14 dígitos; se vier com 11–13 (zero à esquerda perdido) faz zfill(14).
    CPF (11 dígitos) é preservado.
    Retorna string vazia se entrada vazia/NaN.
    """
    if doc is None:
        return ""
    s = str(doc).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if len(digits) == 11:
        return digits
    if 12 <= len(digits) <= 14:
        return digits.zfill(14)
    return digits


def _parse_valor(v) -> Optional[float]:
    """Converte VALOR do relatório GCOM em float.

    O GCOM tem um padrão peculiar:
      - Com vírgula -> formato BR (ex.: '1.234,56' ou '1234,56').
      - SEM vírgula nem ponto -> está em CENTAVOS implícitos
        (ex.: '58824' = 588,24 ; '90540' = 905,40).
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if pd.isna(v):
            return None
        # Inteiro vindo do pandas para uma célula sem separador também é centavos
        if float(v).is_integer():
            return float(v) / 100.0
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    neg = s.startswith("-")
    s = re.sub(r"[^\d,.\-]", "", s)
    if neg and not s.startswith("-"):
        s = "-" + s
    if not s or s in ("-", ",", "."):
        return None

    has_comma = "," in s
    has_dot = "." in s

    try:
        if has_comma and has_dot:
            # Formato BR completo: '.' = milhar, ',' = decimal
            return float(s.replace(".", "").replace(",", "."))
        if has_comma:
            # 'XXXX,YY' -> XXXX.YY
            return float(s.replace(",", "."))
        if has_dot:
            # Raro no GCOM: assume formato US ('1234.56')
            return float(s)
        # Sem separador: centavos implícitos
        return int(s) / 100.0
    except ValueError:
        return None


def _norm(s: str) -> str:
    """Normaliza string: maiúsculas, sem acento básico, espaços colapsados."""
    if s is None:
        return ""
    s = str(s).upper()
    s = (
        s.replace("Ç", "C").replace("Á", "A").replace("Â", "A").replace("Ã", "A")
         .replace("É", "E").replace("Ê", "E").replace("Í", "I").replace("Ó", "O")
         .replace("Ô", "O").replace("Õ", "O").replace("Ú", "U").replace("Ü", "U")
         .replace("Nº", "N").replace("º", "").replace("ª", "")
    )
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _looks_like_header(values) -> bool:
    if not values:
        return False
    flat = " | ".join(_norm(v) for v in values)
    hits = sum(1 for sig in HEADER_SIGNATURES if _norm(sig) in flat)
    return hits >= 2


def _find_col(columns, *candidates) -> Optional[str]:
    """Acha no DataFrame a coluna cujo nome normalizado bate com um dos candidatos."""
    norm_map = {_norm(c): c for c in columns}
    for cand in candidates:
        nc = _norm(cand)
        if nc in norm_map:
            return norm_map[nc]
    # Fallback: contém o termo
    for cand in candidates:
        nc = _norm(cand)
        for k, original in norm_map.items():
            if nc and nc in k:
                return original
    return None


def _parse_via_pandas(html: str) -> Optional[pd.DataFrame]:
    """Tenta pandas.read_html — costuma lidar com colspan/rowspan."""
    try:
        tables = pd.read_html(io.StringIO(html), header=0, flavor="lxml")
    except Exception:
        try:
            tables = pd.read_html(io.StringIO(html), header=0, flavor="bs4")
        except Exception:
            return None
    if not tables:
        return None

    best = None
    best_hits = 0
    for df in tables:
        cols = [str(c) for c in df.columns]
        hits = sum(1 for sig in HEADER_SIGNATURES if _norm(sig) in " | ".join(_norm(c) for c in cols))
        if hits > best_hits and len(df) > 0:
            best = df
            best_hits = hits

    if best is None or best_hits < 2:
        return None
    return best


def _parse_via_bs4(html: str) -> pd.DataFrame:
    """Fallback: procura a tabela com cabeçalhos esperados, tolera rows desiguais."""
    soup = BeautifulSoup(html, "lxml") if _has_lxml() else BeautifulSoup(html, "html.parser")

    tables = soup.find_all("table")
    if not tables:
        return pd.DataFrame()

    target_table = None
    header_row_idx = -1
    headers = []
    for table in tables:
        rows = table.find_all("tr")
        for idx, row in enumerate(rows[:10]):  # cabeçalho costuma estar nas primeiras 10 linhas
            cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
            if _looks_like_header(cells):
                target_table = table
                header_row_idx = idx
                headers = cells
                break
        if target_table is not None:
            break

    if target_table is None:
        # Último recurso: maior tabela, cabeçalho na primeira linha
        target_table = max(tables, key=lambda t: len(t.find_all("tr")))
        rows = target_table.find_all("tr")
        if not rows:
            return pd.DataFrame()
        headers = [td.get_text(strip=True) for td in rows[0].find_all(["td", "th"])]
        header_row_idx = 0

    rows = target_table.find_all("tr")
    data = []
    n = len(headers)
    for row in rows[header_row_idx + 1:]:
        cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
        if not cells:
            continue
        first = _norm(cells[0])
        if first.startswith("TOTAL") or first.startswith("SUBTOTAL"):
            continue
        # Tolera tamanhos diferentes
        if len(cells) < n:
            cells = cells + [""] * (n - len(cells))
        elif len(cells) > n:
            cells = cells[:n]
        # Linha completamente vazia? pula
        if not any(c.strip() for c in cells):
            continue
        data.append(cells)

    return pd.DataFrame(data, columns=headers)


def _has_lxml() -> bool:
    try:
        import lxml  # noqa: F401
        return True
    except ImportError:
        return False


def _parse_relatorio(xls_bytes: bytes) -> pd.DataFrame:
    """O .xls do GCOM é HTML disfarçado. Tenta pandas.read_html, cai para BS4."""
    # Heurística de encoding: o XLS do GCOM costuma vir em latin-1, mas se
    # houver BOM/charset utf-8 declarado, respeita.
    raw = xls_bytes
    if raw[:3] == b"\xef\xbb\xbf":
        html = raw.decode("utf-8", errors="replace")
    else:
        html = raw.decode("latin-1", errors="replace")

    df = _parse_via_pandas(html)
    if df is not None and not df.empty:
        # Remove linhas TOTAL/SUBTOTAL caso pandas tenha incluído
        first_col = df.columns[0]
        df = df[~df[first_col].astype(str).str.strip().str.upper().str.startswith(("TOTAL", "SUBTOTAL"))]
        df = df.reset_index(drop=True)
        return df

    return _parse_via_bs4(html)


def _localizar_coluna(ws, *nomes_possiveis: str) -> int | None:
    """Procura na linha 1 o cabeçalho que bata (case-insensitive, trim).
    Retorna o número da coluna (1-based) ou None."""
    alvo = {n.strip().lower() for n in nomes_possiveis}
    for cell in ws[1]:
        if cell.value is None:
            continue
        if str(cell.value).strip().lower() in alvo:
            return cell.column
    return None


def gerar_template(xls_bytes: bytes) -> bytes:
    """Recebe o XLS bruto do GCOM e devolve o template .xlsx preenchido."""
    df = _parse_relatorio(xls_bytes)

    wb = load_workbook(TEMPLATE_PATH)
    if "Preencher" not in wb.sheetnames:
        raise ValueError("Aba 'Preencher' não encontrada no template.")
    ws = wb["Preencher"]

    col_grupo = _localizar_coluna(ws, "Grupo de Pagamento")
    col_valor_grupo = _localizar_coluna(ws, "Valor do Grupo de Pagamento")
    col_fluxo = _localizar_coluna(
        ws,
        "Codigo do fluxo de caixa",
        "Código do fluxo de caixa",
    )

    # Limpa quaisquer linhas pré-existentes a partir da linha 2 (o template vem com exemplos).
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Resolve nomes reais das colunas no DataFrame (case/acento-insensitive).
    col_cnpj = _find_col(df.columns, "CNPJ/CPF", "CNPJ", "CPF")
    col_docto = _find_col(df.columns, "Nº DOCTO", "N DOCTO", "NUM DOCTO", "NUMERO DOCTO")
    col_vencto = _find_col(df.columns, "VENCTO", "VENCIMENTO")
    col_entrada = _find_col(df.columns, "DATA DA ENTRADA", "DT ENTRADA", "DATA ENTRADA")
    col_valor = _find_col(df.columns, "VALOR", "VL TITULO", "VL. TITULO")

    for i, row_data in df.iterrows():
        excel_row = i + 2

        for col_num, valor in COLUNAS_FIXAS.items():
            ws.cell(row=excel_row, column=col_num, value=valor)

        if col_grupo:
            ws.cell(row=excel_row, column=col_grupo, value=VALOR_GRUPO_PAGAMENTO)
        if col_fluxo:
            ws.cell(row=excel_row, column=col_fluxo, value=VALOR_FLUXO_CAIXA)

        cnpj_raw = row_data[col_cnpj] if col_cnpj else ""
        n_docto = row_data[col_docto] if col_docto else ""
        vencto = row_data[col_vencto] if col_vencto else ""
        entrada = row_data[col_entrada] if col_entrada else ""
        valor_raw = row_data[col_valor] if col_valor else None

        # Trata NaN do pandas
        if pd.isna(cnpj_raw): cnpj_raw = ""
        if pd.isna(n_docto): n_docto = ""
        if pd.isna(vencto): vencto = ""
        if pd.isna(entrada): entrada = ""

        # CNPJ/CPF: só dígitos, célula como texto pra preservar zero à esquerda
        cnpj_limpo = _limpar_documento(cnpj_raw)
        cell_cnpj = ws.cell(row=excel_row, column=2, value=cnpj_limpo)
        cell_cnpj.number_format = "@"

        ws.cell(row=excel_row, column=3,  value=str(n_docto).strip())
        ws.cell(row=excel_row, column=4,  value=str(n_docto).strip())
        ws.cell(row=excel_row, column=9,  value=_formatar_data(entrada))
        ws.cell(row=excel_row, column=10, value=_formatar_data(vencto))
        ws.cell(row=excel_row, column=11, value=_formatar_data(vencto))

        # Valor do Grupo de Pagamento (vem da coluna VALOR do relatório)
        if col_valor_grupo:
            valor_num = _parse_valor(valor_raw)
            if valor_num is not None:
                cell_valor = ws.cell(row=excel_row, column=col_valor_grupo, value=valor_num)
                cell_valor.number_format = '#,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
